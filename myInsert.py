import re
import openpyxl as pyxl


def insert(sql, dbname):
    if 'values' in sql:
        matchObj = re.search(r'^insert into (.*)\((.*)\) values\((.*)\);$', sql)
        if matchObj:
            table_name = matchObj.group(1)
            col = matchObj.group(2).split(',')
            values = matchObj.group(3).split(',')
            # c_v = dict(zip(col, values))
            # keys = list(c_v.keys())
            # values = list(c_v.values())
            db = pyxl.load_workbook("data/" + dbname + ".xlsx")
            tb = db[table_name]
            tb_row = tb.max_row
            for i in range(0, len(col)):
                for j in range(1, tb.max_column + 1):
                    if tb.cell(1, j).value == col[i]:
                        tb.cell(tb_row + 1, j).value = values[i]
            db.save("data/data02.xlsx")
            db.close()
            print('Insert success')
        else:
            print('SQL error')
    else:
        print('SQL error, Please check and try again')


# sql = 'insert into sk(Name,Num,Age,Class) values(wcl,5,21,1);'

# insert(sql, 'data02')
