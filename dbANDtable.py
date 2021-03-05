import openpyxl as pyxl

db_path = 'data/'


def create_db(dbname):
    dbpath = 'data/' + dbname + '.xlsx'
    database = pyxl.Workbook()
    database.save(dbpath)
    create_tb(dbname)
    print("db_create successfully")


def create_tb(dbname):
    db = pyxl.load_workbook("data/data02.xlsx")
    table = db.create_sheet(dbname)
    columns_name = ['table', 'name', 'type', 'null', 'unique', 'primary_key', 'foreign_key']
    for i in range(len(columns_name)):
        table.cell(row=1, column=i + 1).value = columns_name[i]
    if db.worksheets[0].title == 'Sheet':
        del db['Sheet']
    db.save("data/data02.xlsx")
    db.close()


def create_table(table_name):
    current_database = pyxl.load_workbook('data/data02.xlsx')
    if table_name not in current_database.sheetnames:
        table = current_database.create_sheet(table_name)
        columns_name = ['Name', 'Num', 'Age', 'Class']
        for i in range(len(columns_name)):
            table.cell(row=1, column=i + 1).value = columns_name[i]
        current_database.save("data/data02.xlsx")
    else:
        print("table exist")
        return
    print("table_create successfully")