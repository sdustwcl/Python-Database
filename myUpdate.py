import re
import openpyxl as pyxl


def update(sql, dbname):
    if 'set' in sql:
        if 'where' in sql:
            matchObj = re.search(r'^update (.*) set (.*) where (.*);$', sql)
            table_name = matchObj.group(1)
            set_obj = matchObj.group(2).split('=')
            pos = matchObj.group(3).split('=')
            db = pyxl.load_workbook("data/" + dbname + ".xlsx")
            tb = db[table_name]
            tb_row = tb.max_row
            tb_col = tb.max_column
            x = 0
            for i in range(1, tb_col + 1):
                if set_obj[0] == tb.cell(1, i).value:
                    x = i
            for m in range(1, tb_col + 1):
                for n in range(1, tb_row + 1):
                    if pos[1] == str(tb.cell(n, m).value):
                        tb.cell(n, x).value = set_obj[1]
            db.save("data/" + dbname + ".xlsx")
            db.close()
            print('Update success')
        else:
            matchObj = re.search(r'^update (.*) set (.*);$', sql)
            table_name = matchObj.group(1)
            obj = matchObj.group(2).split('=')
            db = pyxl.load_workbook("data/" + dbname + ".xlsx")
            tb = db[table_name]
            tb_row = tb.max_row
            tb_col = tb.max_column
            if '+' in sql:
                val = obj[1].split('+')
                for i in range(1, tb_col + 1):
                    if obj[0] == tb.cell(1, i).value:
                        y = i
                for j in range(2, tb_row + 1):
                    n = int(tb.cell(j, y).value)
                    n += int(val[1])
                    tb.cell(j, y).value = str(n)
                db.save("data/" + dbname + ".xlsx")
                db.close()
                print('Update success')
            elif '-' in sql:
                val = obj[1].split('-')
                for i in range(1, tb_col + 1):
                    if obj[0] == tb.cell(1, i).value:
                        y = i
                for j in range(2, tb_row + 1):
                    n = int(tb.cell(j, y).value)
                    n -= int(val[1])
                    tb.cell(j, y).value = str(n)
                db.save("data/" + dbname + ".xlsx")
                db.close()
                print('Update success')
            elif '*' in sql:
                val = obj[1].split('*')
                for i in range(1, tb_col + 1):
                    if obj[0] == tb.cell(1, i).value:
                        y = i
                for j in range(2, tb_row + 1):
                    n = int(tb.cell(j, y).value)
                    n *= int(val[1])
                    tb.cell(j, y).value = str(n)
                db.save("data/" + dbname + ".xlsx")
                db.close()
                print('Update success')
            elif '/' in sql:
                val = obj[1].split('/')
                for i in range(1, tb_col + 1):
                    if obj[0] == tb.cell(1, i).value:
                        y = i
                for j in range(2, tb_row + 1):
                    n = int(tb.cell(j, y).value)
                    n /= int(val[1])
                    tb.cell(j, y).value = str(n)
                db.save("data/" + dbname + ".xlsx")
                db.close()
                print('Update success')
            else:
                print('ERROR')
    else:
        print('SQL error, Please check and try again')

# sql = 'update sk set Age=Age+1;'
# sql = 'update sk set Age=Age-1;'
# sql = 'update sk set Age=23 where Class=1;'
# update(sql, 'data02')
